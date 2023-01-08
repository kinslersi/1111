import requests
from bs4 import BeautifulSoup
import time
import openpyxl


def pipeline(salary):
    head=salary[:2]
    price=""
    for i in salary:
        if i=="萬" or i=="~" or i=="." or i.isdigit():
            price+=i
        else:
            pass
    if "萬" in price:
        price=price.replace("萬","")
        if "~" in price:
            low_price,high_price=price.split("~")
            low_price=float(low_price)*10000
            high_price=float(high_price)*10000
            mean_price=(low_price+high_price)/2
        else:
            low_price=float(price)*10000
            high_price=float(price)*10000
            mean_price=float(price)*10000
    else:
        if "~" in price:
            low_price,high_price=price.split("~")
            mean_price=(float(low_price)+float(high_price))/2
        else:
            low_price=price
            high_price=price
            mean_price=(float(low_price)+float(high_price))/2
    return head,low_price,high_price,mean_price

wb=openpyxl.Workbook()
ws=wb.active
ws['A1']='職缺名稱'     # 指定excel儲存格，不然會亂掉
ws['B1']='公司名稱'
ws['C1']='職缺連結'
ws['D1']='職缺地區'
ws['E1']='薪資待遇'
ws['F1']='計薪方式'
ws['G1']='最低薪資'
ws['H1']='最高薪資'
ws['I1']='平均薪資'


page=0
while True:
    page+=1
    print('page :',page)
    res=requests.get(f'https://www.1111.com.tw/search/job?ks=%E5%A4%A7%E6%95%B8%E6%93%9A&page={page}')
    soup=BeautifulSoup(res.text,'html.parser')
    if soup.find_all('div',class_="job_item_info")==[]:
        print('done')
        break
    for one_line in soup.find_all('div',class_="job_item_info"):
        name=one_line.find('h5',class_="card-title title_6").text
        url=one_line.find('a')['href']
        company=one_line.find('h6',class_="job_item_company mb-1 digit_5 body_3").text
        place=one_line.find('a',class_="job_item_detail_location mr-3 position-relative").text
        salary=one_line.find('div',class_="job_item_detail_salary ml-3 font-weight-style digit_6").text
        head,low_price,high_price,mean_price=pipeline(salary)
        ws.append([name,company,url,place,salary,head,low_price,high_price,mean_price])    # 以row為單位加入openpyxl worksheet
    time.sleep(2)
wb.save('1111.xlsx')



